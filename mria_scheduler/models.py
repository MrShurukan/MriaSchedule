from __future__ import annotations

from dataclasses import dataclass
from re import sub

from openpyxl.styles import PatternFill


def normalize_text(value: object) -> str:
    """Normalize arbitrary cell/input text for matching."""
    if value is None:
        return ""
    text = str(value).strip().lower()
    # Collapse multiple spaces to improve exact/fuzzy stability.
    text = sub(r"\s+", " ", text)
    return text


def cache_key(partner_name: str, event_name: str) -> str:
    return f"{normalize_text(partner_name)}|||{normalize_text(event_name)}"


@dataclass(slots=True)
class PartnerRecord:
    row_index: int
    event_name: str
    partner_name: str
    technical_requirements: str

    @property
    def event_name_norm(self) -> str:
        return normalize_text(self.event_name)

    @property
    def partner_name_norm(self) -> str:
        return normalize_text(self.partner_name)


@dataclass(slots=True)
class DayShiftColumn:
    day_label: str
    shift_label: str
    column_index: int


@dataclass(slots=True)
class DistributionEvent:
    day_label: str
    shift_label: str
    partner_name: str
    event_name: str
    color_key: str
    fill: PatternFill
    row_index: int
    column_index: int


@dataclass(slots=True)
class OutputScheduleEvent:
    day_label: str
    shift_label: str
    event_name: str
    location: str | None
    partner_name: str
    technical_requirements: str
    fill: PatternFill
