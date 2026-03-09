from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .models import DayShiftColumn, DistributionEvent, PartnerRecord, normalize_text


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text


def _shift_label(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _cell_text(value)


def _effective_cell_value(sheet: Worksheet, row: int, col: int) -> object:
    cell = sheet.cell(row=row, column=col)
    if cell.value not in (None, ""):
        return cell.value
    for merged in sheet.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return sheet.cell(row=merged.min_row, column=merged.min_col).value
    return cell.value


def extract_color_key(fill: PatternFill) -> str:
    if fill is None:
        return ""
    if fill.patternType in (None, "none"):
        return ""
    color = fill.fgColor
    if color is None:
        return ""

    if color.type == "rgb" and color.rgb:
        return f"rgb:{str(color.rgb).upper()}"
    if color.type == "theme" and color.theme is not None:
        tint = color.tint if color.tint is not None else 0
        return f"theme:{color.theme}:{tint}"
    if color.type == "indexed" and color.indexed is not None:
        return f"indexed:{color.indexed}"
    if color.rgb:
        return f"rgb:{str(color.rgb).upper()}"
    if color.indexed is not None:
        return f"indexed:{color.indexed}"
    return ""


def load_partners_records(path: Path, logger) -> list[PartnerRecord]:
    logger.info("Загрузка файла 'Партнёры': %s", path)
    workbook = load_workbook(path, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        logger.info("Лист партнёров: '%s'", sheet.title)

        header_to_col: dict[str, int] = {}
        for col in range(1, sheet.max_column + 1):
            header_text = _cell_text(sheet.cell(row=1, column=col).value)
            if not header_text:
                continue
            header_to_col[normalize_text(header_text)] = col

        event_col = header_to_col.get(normalize_text("Название мастер-класса"))
        partner_col = header_to_col.get(normalize_text("Название организации"))
        tz_col = header_to_col.get(normalize_text("Оборудование ТЗ"))
        if not event_col or not partner_col or not tz_col:
            raise ValueError(
                "В файле 'Партнёры' не найдены обязательные столбцы: "
                "'Название мастер-класса', 'Название организации', 'Оборудование ТЗ'"
            )

        records: list[PartnerRecord] = []
        for row in range(2, sheet.max_row + 1):
            event_name = _cell_text(sheet.cell(row=row, column=event_col).value)
            partner_name = _cell_text(sheet.cell(row=row, column=partner_col).value)
            if not event_name and not partner_name:
                continue
            technical_requirements = _cell_text(sheet.cell(row=row, column=tz_col).value)
            records.append(
                PartnerRecord(
                    row_index=row,
                    event_name=event_name,
                    partner_name=partner_name,
                    technical_requirements=technical_requirements,
                )
            )

        logger.info("Загружено строк партнёров: %d", len(records))
        return records
    finally:
        workbook.close()


def parse_distribution_workbook(
    path: Path, logger
) -> tuple[dict[str, str], list[DayShiftColumn], list[DistributionEvent], bytes | None]:
    logger.info("Загрузка файла 'Распределение': %s", path)
    workbook = load_workbook(path, data_only=False)
    try:
        if len(workbook.worksheets) < 2:
            raise ValueError("Файл 'Распределение' должен содержать минимум 2 листа")

        legend_sheet = workbook.worksheets[0]
        program_sheet = workbook.worksheets[1]
        logger.info("Лист условных обозначений: '%s'", legend_sheet.title)
        logger.info("Лист программы: '%s'", program_sheet.title)

        color_to_location: dict[str, str] = {}
        for row in range(1, legend_sheet.max_row + 1):
            legend_cell = legend_sheet.cell(row=row, column=3)
            location_name = _cell_text(legend_cell.value)
            if not location_name:
                continue
            color_key = extract_color_key(legend_cell.fill)
            if not color_key:
                logger.warning(
                    "Пропуск строки %d на листе '%s': у локации '%s' нет цвета",
                    row,
                    legend_sheet.title,
                    location_name,
                )
                continue
            previous_location = color_to_location.get(color_key)
            if previous_location and previous_location != location_name:
                logger.warning(
                    "Цвет '%s' уже назначен локации '%s', новое значение '%s' будет проигнорировано",
                    color_key,
                    previous_location,
                    location_name,
                )
                continue
            color_to_location[color_key] = location_name

        if not color_to_location:
            raise ValueError(
                "Не удалось построить справочник локаций: на листе 'условные обозначения' "
                "в колонке C нет непустых цветных ячеек"
            )
        logger.info("Построен справочник цветов: %d локаций", len(color_to_location))

        day_shift_columns: list[DayShiftColumn] = []
        previous_day = ""
        for col in range(2, program_sheet.max_column + 1):
            day_raw = _effective_cell_value(program_sheet, row=1, col=col)
            shift_raw = _effective_cell_value(program_sheet, row=2, col=col)

            day_label = _cell_text(day_raw)
            if day_label:
                previous_day = day_label
            elif previous_day:
                day_label = previous_day

            shift_label = _shift_label(shift_raw)
            if not shift_label:
                continue
            if not day_label:
                logger.warning("Столбец %d имеет смену '%s', но день не определён. Пропускаю.", col, shift_label)
                continue

            day_shift_columns.append(
                DayShiftColumn(
                    day_label=day_label,
                    shift_label=shift_label,
                    column_index=col,
                )
            )

        if not day_shift_columns:
            raise ValueError("Не удалось найти столбцы смен на листе 'программа'")

        unique_days = list(dict.fromkeys(item.day_label for item in day_shift_columns))
        logger.info("Найдено дней: %d; найдено смен-столбцов: %d", len(unique_days), len(day_shift_columns))

        events: list[DistributionEvent] = []
        max_row = program_sheet.max_row
        for column_info in day_shift_columns:
            logger.info(
                "Сканирование дня '%s', смена '%s', столбец %d",
                column_info.day_label,
                column_info.shift_label,
                column_info.column_index,
            )
            for row in range(3, max_row + 1):
                partner_name = _cell_text(program_sheet.cell(row=row, column=1).value)
                if not partner_name:
                    continue

                event_cell = program_sheet.cell(row=row, column=column_info.column_index)
                event_name = _cell_text(event_cell.value)
                if not event_name:
                    continue

                color_key = extract_color_key(event_cell.fill)
                events.append(
                    DistributionEvent(
                        day_label=column_info.day_label,
                        shift_label=column_info.shift_label,
                        partner_name=partner_name,
                        event_name=event_name,
                        color_key=color_key,
                        fill=copy(event_cell.fill),
                        row_index=row,
                        column_index=column_info.column_index,
                    )
                )

        logger.info("Найдено заполненных ячеек мероприятий: %d", len(events))
        return color_to_location, day_shift_columns, events, workbook.loaded_theme
    finally:
        workbook.close()
