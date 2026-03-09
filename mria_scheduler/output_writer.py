from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from .models import OutputScheduleEvent


def default_output_filename(now: datetime | None = None) -> str:
    ts = now or datetime.now()
    return (
        f"Расписание_{ts.day:02d}_{ts.month:02d}_{ts.year}_"
        f"{ts.hour:02d}_{ts.minute:02d}_{ts.second:02d}.xlsx"
    )


def resolve_output_path(cwd: Path, custom_name: str | None) -> Path:
    filename = custom_name.strip() if custom_name else default_output_filename()
    output_path = Path(filename)
    if not output_path.suffix:
        output_path = output_path.with_suffix(".xlsx")
    if not output_path.is_absolute():
        output_path = (cwd / output_path).resolve()
    if not output_path.parent.exists():
        raise ValueError(f"Директория для выходного файла не существует: {output_path.parent}")
    return output_path


def _write_separator_row(sheet, row_index: int, text: str) -> None:
    sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=4)
    cell = sheet.cell(row=row_index, column=1, value=text)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def write_schedule_workbook(
    path: Path, events: list[OutputScheduleEvent], logger, source_theme: bytes | None = None
) -> None:
    logger.info("Формирую выходной файл: %s", path)
    workbook = Workbook()
    if source_theme:
        workbook.loaded_theme = source_theme
    sheet = workbook.active
    sheet.title = "Расписание"

    headers = ["Название мероприятия", "Локация", "Кто проводит", "ТЗ"]
    for col, header in enumerate(headers, start=1):
        header_cell = sheet.cell(row=1, column=col, value=header)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 48
    sheet.column_dimensions["B"].width = 24
    sheet.column_dimensions["C"].width = 35
    sheet.column_dimensions["D"].width = 44

    current_day = ""
    current_shift = ""
    row_index = 1
    for event in events:
        if event.day_label != current_day:
            row_index += 1
            _write_separator_row(sheet, row_index, event.day_label)
            current_day = event.day_label
            current_shift = ""

        if event.shift_label != current_shift:
            row_index += 1
            _write_separator_row(sheet, row_index, f"{event.shift_label} смена")
            current_shift = event.shift_label

        row_index += 1
        sheet.cell(row=row_index, column=1, value=event.event_name)
        sheet.cell(row=row_index, column=2, value=event.location if event.location else None)
        sheet.cell(row=row_index, column=3, value=event.partner_name)
        sheet.cell(row=row_index, column=4, value=event.technical_requirements)

        source_cell = sheet.cell(row=row_index, column=1)
        source_cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.cell(row=row_index, column=2).alignment = Alignment(vertical="top", wrap_text=True)
        sheet.cell(row=row_index, column=3).alignment = Alignment(vertical="top", wrap_text=True)
        sheet.cell(row=row_index, column=4).alignment = Alignment(vertical="top", wrap_text=True)

        if event.fill and event.fill.patternType not in (None, "none"):
            source_cell.fill = copy(event.fill)

    workbook.save(path)
    workbook.close()
    logger.info("Готово. Записано строк данных: %d", len(events))
